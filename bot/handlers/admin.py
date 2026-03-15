# ============================================
# 👑 Admin Handler - Telegram Giveaway Bot
# 📁 File: bot/handlers/admin.py
# 👤 Created by: User with AI
# 📝 Handles the secret /shep admin command with password
#    authentication. Provides admin panel with stats, text
#    editing, Excel export, broadcast, and channel settings.
#    This command is HIDDEN and undocumented by design.
# 📅 Created: 2026-03-15 07:47 (Tashkent Time)
# ============================================
# 📋 CHANGE LOG:
# [2026-03-15 08:26 Tashkent] - Added full channel management
# [2026-03-15 09:18 Tashkent] - Fixed /shep welcome text to 'Xush kelibsiz, Shep!'
#   Admin reply keyboard shown after login. Fixed admin_close action.
# [2026-03-15 09:28 Tashkent] - Broadcast upgraded to use `copy_message`,
#   supporting all media types (video, document, etc).
#   Added cache clearing when admin edits text.
# ============================================

import os
import io
import asyncio
import logging
from datetime import datetime, timezone, timedelta

from aiogram import Router, F, Bot
from aiogram.types import (
    Message, CallbackQuery, BufferedInputFile
)
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from bot import database as db
from bot.code_generator import generate_unique_code
from bot.config import config
from bot.texts.messages import get_message, clear_text_cache
from bot.keyboards.keyboards import (
    get_admin_keyboard,
    get_text_keys_keyboard,
    get_language_select_keyboard,
    get_confirm_broadcast_keyboard,
    get_main_menu_keyboard,
    get_admin_menu_keyboard,
)

logger = logging.getLogger(__name__)


# 🔄 FSM States for admin actions
class AdminStates(StatesGroup):
    waiting_password = State()      # 🔐 Waiting for admin password
    editing_text = State()          # 📝 Waiting for new text content
    broadcast_compose = State()     # 📢 Composing broadcast message
    broadcast_confirm = State()     # ✅ Confirming broadcast
    # 📢 Channel management states
    ch_waiting_type = State()       # 📢 Waiting for channel type
    ch_waiting_url = State()        # 🔗 Waiting for channel URL
    ch_waiting_name = State()       # 📛 Waiting for channel name
    ch_waiting_id = State()         # 🆔 Waiting for Telegram channel ID


# 🛤️ Router
router = Router()

# 📨 Store last broadcast message IDs for deletion
_last_broadcast = {"messages": []}  # [{"chat_id": ..., "message_id": ...}, ...]


# =============================================
# 🧹 EMERGENCY /TOZALA COMMAND (Brute-force delete recent bot messages)
# =============================================

@router.message(Command("tozala"))
async def cmd_cleanup(message: Message, bot: Bot, lang: str = "uz"):
    """🧹 Emergency: try to delete the last N messages bot sent to all users"""
    if not await db.is_user_admin(message.from_user.id):
        return

    user_ids = await db.get_all_user_ids()
    total_deleted = 0
    total_failed = 0

    status_msg = await message.answer(
        f"🧹 <b>Tozalash boshlandi...</b>\n"
        f"Har bir userning oxirgi 50 ta xabar ID sini sinab ko'raman.\n"
        f"Users: {len(user_ids)}",
        parse_mode="HTML"
    )

    for uid in user_ids:
        # 🔍 Try message IDs in a wide recent range
        # We try from a high number downwards
        for offset in range(1, 51):
            try:
                # Try to send a dummy message to find approximately the latest message ID
                # then delete backwards from there
                pass
            except Exception:
                pass

        # Alternative: try deleting by message_id directly
        # Telegram private chat message IDs are relatively small sequential numbers
        # We try a smart range based on a probe
        try:
            probe = await bot.send_message(uid, ".")
            latest_id = probe.message_id
            await bot.delete_message(uid, latest_id)  # Delete the probe

            # Now try deleting backwards from the latest ID
            for msg_id in range(latest_id - 1, max(latest_id - 51, 0), -1):
                try:
                    await bot.delete_message(uid, msg_id)
                    total_deleted += 1
                except Exception:
                    total_failed += 1

        except Exception as e:
            logger.debug(f"🧹 Probe failed for {uid}: {e}")

        await asyncio.sleep(0.5)  # Rate limit

    await status_msg.edit_text(
        f"✅ <b>Tozalash tugadi!</b>\n\n"
        f"🗑️ O'chirildi: {total_deleted}\n"
        f"❌ O'chirib bo'lmadi: {total_failed}\n\n"
        f"<i>Eslatma: Bot faqat o'zi yuborgan xabarlarni o'chira oladi.</i>",
        parse_mode="HTML"
    )


# =============================================
# 🔐 SECRET /SHEP COMMAND (HIDDEN)
# =============================================

@router.message(Command("shep"))
async def cmd_shep(message: Message, state: FSMContext, lang: str = "uz"):
    """🔐 Secret admin entry point — HIDDEN, UNDOCUMENTED"""
    
    # 🛑 Require registration first
    user = await db.get_user(message.from_user.id)
    if not user:
        await message.answer("⚠️ Iltimos, oldin /start buyrug'ini yuborib ro'yxatdan o'ting.")
        return

    # 🔍 Check if user is already admin
    is_admin = await db.is_user_admin(message.from_user.id)

    if is_admin:
        # 👑 Already admin, show admin reply keyboard + inline panel
        await message.answer(
            "👑 <b>Xush kelibsiz, Shep!</b>",
            parse_mode="HTML",
            reply_markup=get_admin_menu_keyboard(lang)
        )
        return

    # 🔐 Ask for password
    text = await get_message("enter_password", lang)
    await message.answer(text, parse_mode="HTML")
    await state.set_state(AdminStates.waiting_password)


@router.message(AdminStates.waiting_password)
async def on_password_entered(message: Message, state: FSMContext, lang: str = "uz"):
    """🔐 Verify admin password"""
    if message.text == config.ADMIN_PASSWORD:
        # ✅ Correct password → make admin
        # Save user to DB first just in case they skipped /start registration
        await db.save_user(
            telegram_id=message.from_user.id,
            username=message.from_user.username,
            first_name=message.from_user.first_name,
            last_name=message.from_user.last_name,
            language=lang
        )
        await db.set_admin(message.from_user.id, True)
        logger.info(f"👑 New admin: {message.from_user.id} ({message.from_user.first_name})")

        # 👑 Show admin reply keyboard
        await message.answer(
            "👑 <b>Xush kelibsiz, Shep!</b>",
            parse_mode="HTML",
            reply_markup=get_admin_menu_keyboard(lang)
        )
        await state.clear()
    else:
        # ❌ Wrong password
        text = await get_message("wrong_password", lang)
        await message.answer(text, parse_mode="HTML")
        await state.clear()


# =============================================
# 📊 STATISTICS
# =============================================

@router.callback_query(F.data == "admin_stats")
async def on_admin_stats(callback: CallbackQuery, lang: str = "uz"):
    """📊 Show bot statistics"""
    # 👑 Verify admin
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    await callback.answer()
    stats = await db.get_statistics()
    text = await get_message("stats_message", lang, **stats)
    await callback.message.edit_text(
        text, parse_mode="HTML",
        reply_markup=get_admin_keyboard(lang)
    )


# =============================================
# 📝 EDIT TEXTS
# =============================================

@router.callback_query(F.data == "admin_edit_texts")
async def on_admin_edit_texts(callback: CallbackQuery, lang: str = "uz"):
    """📝 Show text keys for editing"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    await callback.answer()
    text = await get_message("select_text_key", lang)
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=get_text_keys_keyboard())


@router.callback_query(F.data.startswith("edit_text_"))
async def on_edit_text_key_selected(callback: CallbackQuery):
    """📝 Text key selected, now choose language"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    text_key = callback.data.replace("edit_text_", "")
    await callback.answer()
    await callback.message.edit_text(
        f"🌐 <b>{text_key}</b> — qaysi til uchun o'zgartirasiz?",
        parse_mode="HTML",
        reply_markup=get_language_select_keyboard(text_key)
    )


@router.callback_query(F.data.startswith("textlang_"))
async def on_text_language_selected(callback: CallbackQuery, state: FSMContext, lang: str = "uz"):
    """🌐 Language selected for text editing, show current text and ask for new"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    # 🔍 Parse: textlang_{text_key}_{lang_code}
    parts = callback.data.split("_", 2)  # textlang, key, lang
    if len(parts) < 3:
        await callback.answer("❌ Error", show_alert=True)
        return

    text_key = parts[1]
    target_lang = parts[2]

    # 💾 Save to FSM
    await state.update_data(edit_text_key=text_key, edit_text_lang=target_lang)
    await callback.answer()

    # 📝 Show current text (from DB or fallback)
    current_text = await get_message(text_key, target_lang)

    send_new = await get_message("send_new_text", lang)
    
    # ℹ️ Markdown hint
    md_hint = (
        "✏️ <b>Markdown yordamchi:</b>\n"
        "• *qalin* ➝ yozuvni qalin qiladi\n"
        "• _qiya_ ➝ yozuvni qiya qiladi\n"
        "• `nusxa` ➝ bitta bosishda nusxalanadigan kod\n"
    )

    await callback.message.edit_text(
        f"📝 <b>{text_key}</b> [{target_lang.upper()}]\n\n"
        f"📄 <b>Hozirgi text:</b>\n<code>{current_text}</code>\n\n"
        f"{md_hint}\n"
        f"{'─' * 30}\n"
        f"{send_new}",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.editing_text)


@router.message(AdminStates.editing_text, F.photo)
async def on_new_text_with_photo(message: Message, state: FSMContext, lang: str = "uz"):
    """📝 Admin sent new text with photo"""
    state_data = await state.get_data()
    text_key = state_data.get("edit_text_key")
    target_lang = state_data.get("edit_text_lang")

    if not text_key or not target_lang:
        await state.clear()
        return

    # 🖼️ Get photo file_id (largest size)
    photo_id = message.photo[-1].file_id
    caption = message.caption or ""

    await db.update_text(text_key, target_lang, caption, photo_id)
    clear_text_cache()  # 🧹 Clear text cache

    text = await get_message("text_updated", lang)
    await message.answer(text, parse_mode="HTML", reply_markup=get_admin_keyboard(lang))
    await state.clear()


@router.message(AdminStates.editing_text, F.text)
async def on_new_text(message: Message, state: FSMContext, lang: str = "uz"):
    """📝 Admin sent new text (no photo)"""
    state_data = await state.get_data()
    text_key = state_data.get("edit_text_key")
    target_lang = state_data.get("edit_text_lang")

    if not text_key or not target_lang:
        await state.clear()
        return

    await db.update_text(text_key, target_lang, message.text)
    clear_text_cache()  # 🧹 Clear text cache

    text = await get_message("text_updated", lang)
    await message.answer(text, parse_mode="HTML", reply_markup=get_admin_keyboard(lang))
    await state.clear()


# =============================================
# 📥 EXCEL EXPORT
# =============================================

@router.callback_query(F.data == "admin_export")
async def on_admin_export(callback: CallbackQuery, bot: Bot, lang: str = "uz"):
    """📥 Export all users + codes to formatted Excel"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    await callback.answer()
    gen_text = await get_message("export_generating", lang)
    await callback.message.edit_text(gen_text, parse_mode="HTML")

    try:
        # 📊 Fetch all users
        users = await db.get_all_users_for_export()

        # 📝 Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Giveaway Users"

        # 🎨 Styles
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        # 🎨 Alternating row colors
        light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

        # 📋 Headers
        headers = [
            "№", "Telegram ID", "Username", "Ism",
            "Familiya", "Telefon", "Til", "Tasdiqlangan",
            "Kodlar", "Referal soni", "Jami kodlar", "Ro'yxatdan o'tgan sana"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 📝 Data rows
        for idx, user in enumerate(users, 1):
            row_num = idx + 1
            row_data = [
                idx,
                user.get("telegram_id", ""),
                user.get("username", "") or "—",
                user.get("first_name", "") or "—",
                user.get("last_name", "") or "—",
                user.get("phone", "") or "—",
                user.get("language", "uz").upper(),
                "✅" if user.get("is_verified") else "❌",
                user.get("codes_list", "") or "—",
                user.get("referral_count", 0),
                user.get("total_codes", 0),
                user.get("created_at", "").strftime("%d.%m.%Y %H:%M") if user.get("created_at") else "—"
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col in (1, 7, 8, 10, 11):
                    cell.alignment = center_alignment
                else:
                    cell.alignment = cell_alignment
                # 🎨 Alternating row colors
                if idx % 2 == 0:
                    cell.fill = light_fill

        # 📏 Auto-width columns with extra spacing
        for col in range(1, len(headers) + 1):
            max_length = len(str(headers[col - 1]))
            for row in range(2, len(users) + 2):
                cell_value = str(ws.cell(row=row, column=col).value or "")
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            # ➕ Add padding
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_length + 4, 50)

        # 🔒 Freeze header row
        ws.freeze_panes = "A2"

        # 💾 Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 📅 Current date for filename
        tashkent_tz = timezone(timedelta(hours=5))
        now = datetime.now(tashkent_tz)
        filename = f"giveaway_users_{now.strftime('%Y%m%d_%H%M')}.xlsx"

        # 📤 Send file
        doc = BufferedInputFile(output.read(), filename=filename)
        done_text = await get_message("export_done", lang)
        await bot.send_document(
            callback.from_user.id,
            doc,
            caption=done_text
        )

        logger.info(f"📥 Excel export sent to admin {callback.from_user.id} ({len(users)} users)")

    except Exception as e:
        logger.error(f"❌ Excel export error: {e}")
        await callback.message.answer(f"❌ Export error: {e}")


# =============================================
# 📢 BROADCAST
# =============================================

@router.callback_query(F.data == "admin_broadcast")
async def on_admin_broadcast(callback: CallbackQuery, state: FSMContext, lang: str = "uz"):
    """📢 Start broadcast composition"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    await callback.answer()
    text = await get_message("send_broadcast_text", lang)
    await callback.message.edit_text(text, parse_mode="HTML")
    await state.set_state(AdminStates.broadcast_compose)


@router.message(AdminStates.broadcast_compose)
async def on_broadcast_message(message: Message, state: FSMContext, lang: str = "uz"):
    """📢 Admin composed any kind of broadcast message (text/photo/video/etc)"""
    await state.update_data(
        broadcast_msg_id=message.message_id,
        broadcast_from_chat=message.chat.id
    )

    confirm_text = await get_message("broadcast_confirm", lang, message="")
    
    # 🖼️ Show preview by copying the message
    await message.copy_to(chat_id=message.chat.id)
    await message.answer(
        confirm_text,
        parse_mode="HTML",
        reply_markup=get_confirm_broadcast_keyboard(lang)
    )
    
    await state.set_state(AdminStates.broadcast_confirm)


@router.callback_query(F.data == "broadcast_confirm")
async def on_broadcast_confirmed(callback: CallbackQuery, bot: Bot, state: FSMContext, lang: str = "uz"):
    """✅ Execute broadcast to all users"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    await callback.answer()
    state_data = await state.get_data()
    msg_id = state_data.get("broadcast_msg_id")
    from_chat = state_data.get("broadcast_from_chat")

    if not msg_id or not from_chat:
        await callback.message.answer("❌ Xatolik yuz berdi. Qaytadan urinib ko'ring.")
        await state.clear()
        return

    # 📋 Get all user IDs
    user_ids = await db.get_all_user_ids()
    total = len(user_ids)

    started_text = await get_message("broadcast_started", lang, total=total)
    progress_msg = await callback.message.edit_text(started_text, parse_mode="HTML")

    # 🧹 Clear previous broadcast message IDs
    _last_broadcast["messages"] = []

    sent = 0
    failed = 0
    blocked = 0

    for i, uid in enumerate(user_ids):
        try:
            # 🚀 COPY MESSAGE - Native support for all media types and formatting!
            result = await bot.copy_message(chat_id=uid, from_chat_id=from_chat, message_id=msg_id)
            sent += 1
            # 💾 Store message ID for potential deletion
            _last_broadcast["messages"].append({"chat_id": uid, "message_id": result.message_id})
        except Exception as e:
            error_text = str(e).lower()
            if "blocked" in error_text or "deactivated" in error_text:
                blocked += 1
            else:
                failed += 1
            logger.debug(f"📢 Broadcast failed for {uid}: {e}")

        # ⏱️ Rate limiting: 25 messages/second
        if (i + 1) % config.BROADCAST_RATE == 0:
            await asyncio.sleep(1)

        # 📊 Update progress every 100 messages
        if (i + 1) % 100 == 0:
            try:
                await progress_msg.edit_text(
                    f"📢 Yuborilmoqda... {i+1}/{total}\n"
                    f"📤 Yuborildi: {sent} | ❌ Xato: {failed} | 🚫 Bloklangan: {blocked}"
                )
            except Exception:
                pass

    # ✅ Broadcast complete
    done_text = await get_message("broadcast_done", lang, sent=sent, failed=failed, blocked=blocked)
    
    # 🗑️ Add delete button to the result message
    from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
    done_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🗑️ Oxirgi broadcastni o'chirish", callback_data="delete_last_broadcast")],
        [InlineKeyboardButton(text="👑 Admin panel", callback_data="admin_panel")],
    ])
    await progress_msg.edit_text(done_text, parse_mode="HTML", reply_markup=done_kb)
    await state.clear()

    logger.info(f"📢 Broadcast done: sent={sent}, failed={failed}, blocked={blocked}. Stored {len(_last_broadcast['messages'])} message IDs.")


@router.callback_query(F.data == "broadcast_cancel")
async def on_broadcast_cancelled(callback: CallbackQuery, state: FSMContext, lang: str = "uz"):
    """❌ Cancel broadcast"""
    await callback.answer()
    await state.clear()
    text = await get_message("admin_welcome", lang)
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=get_admin_keyboard(lang))


@router.callback_query(F.data == "delete_last_broadcast")
async def on_delete_broadcast(callback: CallbackQuery, bot: Bot, lang: str = "uz"):
    """🗑️ Delete all messages from the last broadcast"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    messages = _last_broadcast.get("messages", [])
    if not messages:
        await callback.answer("❌ O'chiriladigan xabar topilmadi!", show_alert=True)
        return

    await callback.answer()
    total = len(messages)
    deleted = 0
    failed = 0

    await callback.message.edit_text(
        f"🗑️ O'chirilmoqda... 0/{total}",
        parse_mode="HTML"
    )

    for i, entry in enumerate(messages):
        try:
            await bot.delete_message(chat_id=entry["chat_id"], message_id=entry["message_id"])
            deleted += 1
        except Exception as e:
            failed += 1
            logger.debug(f"🗑️ Delete failed for {entry}: {e}")

        # ⏱️ Rate limiting
        if (i + 1) % 25 == 0:
            await asyncio.sleep(1)

    # 🧹 Clear stored messages
    _last_broadcast["messages"] = []

    await callback.message.edit_text(
        f"✅ <b>Broadcast xabarlari o'chirildi!</b>\n\n"
        f"🗑️ O'chirildi: {deleted}\n"
        f"❌ Xato: {failed}\n"
        f"📊 Jami: {total}",
        parse_mode="HTML",
        reply_markup=get_admin_keyboard(lang)
    )
    logger.info(f"🗑️ Broadcast deleted: {deleted}/{total}")


# =============================================
# 📢 CHANNEL MANAGEMENT (FULL CRUD)
# =============================================

def _build_channels_text_and_keyboard(channels):
    """📢 Build channel list text and inline keyboard"""
    from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup

    text = "📢 <b>Kanal sozlamalari:</b>\n\n"
    buttons = []

    for ch in channels:
        emoji = {"telegram": "📢", "instagram": "📸", "youtube": "▶️"}.get(ch["channel_type"], "🔗")
        ch_id = ch["id"]

        text += f"{emoji} <b>{ch.get('channel_name', '—')}</b> ({ch['channel_type']})\n"
        text += f"   🔗 {ch.get('channel_url', '—')}\n"
        if ch.get("channel_id"):
            text += f"   🆔 {ch['channel_id']}\n"
        text += "\n"

        # 🗑️ Delete button for each channel
        buttons.append([
            InlineKeyboardButton(text=f"{emoji} {ch.get('channel_name', '—')[:20]}", callback_data=f"ch_noop_{ch_id}"),
            InlineKeyboardButton(text="🗑️ O'chirish", callback_data=f"ch_delete_{ch_id}"),
        ])

    # ➕ Add channel button
    buttons.append([InlineKeyboardButton(text="➕ Yangi kanal qo'shish", callback_data="ch_add_new")])
    # 🔙 Back button
    buttons.append([InlineKeyboardButton(text="🔙 Orqaga", callback_data="admin_back")])

    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    return text, keyboard


@router.callback_query(F.data == "admin_channels")
async def on_admin_channels(callback: CallbackQuery, lang: str = "uz"):
    """📢 Show channel management panel"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    await callback.answer()
    channels = await db.get_all_channels()
    text, keyboard = _build_channels_text_and_keyboard(channels)
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=keyboard)





@router.callback_query(F.data.startswith("ch_delete_"))
async def on_channel_delete(callback: CallbackQuery, lang: str = "uz"):
    """🗑️ Delete a channel"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    ch_db_id = int(callback.data.replace("ch_delete_", ""))
    ch = await db.get_channel_by_id(ch_db_id)
    ch_name = ch.get("channel_name", "Unknown") if ch else "Unknown"

    await db.remove_channel(ch_db_id)
    await callback.answer(f"🗑️ {ch_name} o'chirildi!", show_alert=True)
    logger.info(f"🗑️ Channel deleted: {ch_name} (ID: {ch_db_id})")

    # 🔄 Refresh channel list
    channels = await db.get_all_channels()
    text, keyboard = _build_channels_text_and_keyboard(channels)
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=keyboard)


@router.callback_query(F.data.startswith("ch_noop_"))
async def on_channel_noop(callback: CallbackQuery):
    """📢 No-op for channel name button"""
    await callback.answer()


# =============================================
# ➕ ADD NEW CHANNEL FLOW
# =============================================

@router.callback_query(F.data == "ch_add_new")
async def on_add_channel_start(callback: CallbackQuery, state: FSMContext, lang: str = "uz"):
    """➕ Start adding a new channel"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    await callback.answer()
    from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup

    type_keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📢 Telegram", callback_data="ch_type_telegram")],
        [InlineKeyboardButton(text="📸 Instagram", callback_data="ch_type_instagram")],
        [InlineKeyboardButton(text="▶️ YouTube", callback_data="ch_type_youtube")],
        [InlineKeyboardButton(text="🔙 Bekor qilish", callback_data="admin_channels")],
    ])
    await callback.message.edit_text(
        "➕ <b>Yangi kanal qo'shish</b>\n\nKanal turini tanlang:",
        parse_mode="HTML", reply_markup=type_keyboard
    )


@router.callback_query(F.data.startswith("ch_type_"))
async def on_channel_type_selected(callback: CallbackQuery, state: FSMContext, lang: str = "uz"):
    """📢 Channel type selected, ask for URL"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    ch_type = callback.data.replace("ch_type_", "")
    await state.update_data(new_ch_type=ch_type)
    await callback.answer()

    emoji = {"telegram": "📢", "instagram": "📸", "youtube": "▶️"}.get(ch_type, "🔗")
    await callback.message.edit_text(
        f"{emoji} <b>{ch_type.capitalize()}</b> kanal\n\n"
        f"🔗 Kanal URL manzilini yuboring:\n"
        f"<i>(Masalan: https://t.me/channel_name)</i>",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.ch_waiting_url)


@router.message(AdminStates.ch_waiting_url, F.text)
async def on_channel_url_entered(message: Message, state: FSMContext, lang: str = "uz"):
    """🔗 Channel URL entered, ask for name"""
    await state.update_data(new_ch_url=message.text.strip())
    await message.answer(
        "📛 Kanal nomini yuboring:\n<i>(Masalan: My Channel)</i>",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.ch_waiting_name)


@router.message(AdminStates.ch_waiting_name, F.text)
async def on_channel_name_entered(message: Message, state: FSMContext, lang: str = "uz"):
    """📛 Channel name entered, validate and save automatically"""
    await state.update_data(new_ch_name=message.text.strip())
    state_data = await state.get_data()

    if state_data.get("new_ch_type") == "telegram":
        # 🆔 Auto-determine channel ID from URL and validate
        url = state_data.get("new_ch_url", "")
        ch_id = url
        
        # Parse common t.me links to get @username
        if "t.me/" in url:
            parts = url.split("t.me/")
            if len(parts) > 1:
                username = parts[1].split("/")[0].split("?")[0]
                if not username.startswith("+"): # Ignore private invite links
                    ch_id = f"@{username}"

        try:
            # First show a loading message
            loading_msg = await message.answer("⏳ <i>Tekshirilmoqda... Kanal ID aniqlanmoqda.</i>", parse_mode="HTML")
            
            # Validate with Telegram API
            chat = await message.bot.get_chat(ch_id)
            
            # We got the chat successfully, save the exact numeric ID to be safe
            await state.update_data(new_ch_id=str(chat.id))
            await loading_msg.delete()
            await _save_new_channel(message, state, lang)
            
        except Exception as e:
            await message.answer(
                f"❌ <b>Xato: Bunday kanal topilmadi yoki bot u yerda admin emas!</b>\n\n"
                f"Siz kiritgan link orqali kanalni tekshirib bo'lmadi: <code>{ch_id}</code>\n"
                f"Iltimos, botni telegram kanalga admin qilganingizga va to'g'ri ochiq (public) link kiritganingizga ishonch hosil qiling.\n\n"
                f"Xatolik tasfilotlari: {e}",
                parse_mode="HTML"
            )
            # Re-ask for URL
            emoji = "📢"
            await message.answer(
                f"{emoji} <b>Telegram</b> kanal\n\n"
                f"🔗 Kanal URL manzilini yoki username ni yuboring:\n"
                f"<i>(Masalan: https://t.me/channel_name yoki @channel_name)</i>",
                parse_mode="HTML"
            )
            await state.set_state(AdminStates.ch_waiting_url)
    else:
        # 📸/▶️ Instagram/YouTube don't need channel_id validation
        await _save_new_channel(message, state, lang)


async def _save_new_channel(message: Message, state: FSMContext, lang: str = "uz"):
    """💾 Save the new channel to database"""
    state_data = await state.get_data()
    ch_type = state_data.get("new_ch_type", "telegram")
    ch_url = state_data.get("new_ch_url", "")
    ch_name = state_data.get("new_ch_name", "")
    ch_id = state_data.get("new_ch_id")

    await db.add_channel(
        channel_type=ch_type,
        channel_url=ch_url,
        channel_name=ch_name,
        channel_id=ch_id
    )

    logger.info(f"➕ New channel added: {ch_type} - {ch_name} ({ch_url})")

    # ✅ Show success and go back to channel list
    channels = await db.get_all_channels()
    text, keyboard = _build_channels_text_and_keyboard(channels)
    success_text = f"✅ *{ch_name}* kanali qo'shildi!\n\n{text}"
    await message.answer(success_text, parse_mode="HTML", reply_markup=keyboard)
    await state.clear()


# =============================================
# 🔙 ADMIN PANEL NAVIGATION
# =============================================

@router.callback_query(F.data == "admin_back")
async def on_admin_back(callback: CallbackQuery, lang: str = "uz"):
    """🔙 Back to admin panel"""
    if not await db.is_user_admin(callback.from_user.id):
        await callback.answer("⛔", show_alert=True)
        return

    await callback.answer()
    text = await get_message("admin_welcome", lang)
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=get_admin_keyboard(lang))


@router.callback_query(F.data == "admin_close")
async def on_admin_close(callback: CallbackQuery, lang: str = "uz"):
    """🔙 Close admin panel, back to admin reply menu"""
    await callback.answer()
    await callback.message.delete()
    # 👑 Show admin reply keyboard if admin, else user menu
    is_admin = await db.is_user_admin(callback.from_user.id)
    if is_admin:
        await callback.message.answer(
            "👑 Admin panel",
            reply_markup=get_admin_menu_keyboard(lang)
        )
    else:
        await callback.message.answer(
            "📋 Menu",
            reply_markup=get_main_menu_keyboard(lang)
        )
