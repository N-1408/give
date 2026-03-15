# ============================================
# 📋 Menu Handler - Telegram Giveaway Bot
# 📁 File: bot/handlers/menu.py
# 👤 Created by: User with AI
# 📝 Handles all main menu interactions: my chances, referral
#    link, rules, prizes. Also handles admin reply keyboard
#    buttons (stats, export, broadcast, channels, edit texts).
# 📅 Created: 2026-03-15 07:47 (Tashkent Time)
# ============================================
# 📋 CHANGE LOG:
# [2026-03-15 09:18 Tashkent] - Removed top referrers & contact admin
#   handlers. Added admin reply keyboard button handlers (stats, export,
#   broadcast, channels, edit texts).
# ============================================

import logging
from aiogram import Router, F, Bot
from aiogram.types import Message
from aiogram.utils.deep_linking import create_start_link
from aiogram.fsm.context import FSMContext

from bot import database as db
from bot.texts.messages import get_message, get_text_with_image
from bot.keyboards.keyboards import (
    get_admin_keyboard,
    get_text_keys_keyboard,
    get_confirm_broadcast_keyboard,
    get_admin_menu_keyboard,
)

logger = logging.getLogger(__name__)

# 🛤️ Router
router = Router()


# =============================================
# 🎟️ MY CHANCES
# =============================================

@router.message(F.text.in_({
    "🎟️ Mening imkoniyatlarim", "🎟️ Мои шансы", "🎟️ My chances"
}))
async def on_my_chances(message: Message, bot: Bot, lang: str = "uz"):
    """🎟️ Show user's codes and chances count"""
    user_id = message.from_user.id

    # 📋 Get user codes
    codes = await db.get_user_codes(user_id)
    total = len(codes)
    referral_count = sum(1 for c in codes if c["code_type"] == "referral")

    # 📝 Format codes list
    if codes:
        codes_list = "\n".join([
            f"{'🎫' if c['code_type'] == 'subscription' else '🔗'} `{c['code']}` ({c['code_type']})"
            for c in codes
        ])
    else:
        codes_list = "—"

    text = await get_message(
        "my_chances", lang,
        total=total, referral_count=referral_count, codes_list=codes_list
    )
    await message.answer(text, parse_mode="Markdown")


# =============================================
# 🔗 MY REFERRAL LINK
# =============================================

@router.message(F.text.in_({
    "🔗 Referal linkim", "🔗 Моя реферальная ссылка", "🔗 My referral link"
}))
async def on_referral_link(message: Message, bot: Bot, lang: str = "uz"):
    """🔗 Show user's referral link"""
    user_id = message.from_user.id
    referral_link = await create_start_link(bot, f"ref_{user_id}")

    text = await get_message("referral_info", lang, referral_link=referral_link)
    await message.answer(text, parse_mode="Markdown")


# =============================================
# 📋 RULES
# =============================================

@router.message(F.text.in_({
    "📋 Qoidalar", "📋 Правила", "📋 Rules"
}))
async def on_rules(message: Message, lang: str = "uz"):
    """📋 Show giveaway rules"""
    content, image_id = await get_text_with_image("rules", lang)

    if image_id:
        await message.answer_photo(photo=image_id, caption=content, parse_mode="Markdown")
    else:
        await message.answer(content, parse_mode="Markdown")


# =============================================
# 🎁 PRIZES
# =============================================

@router.message(F.text.in_({
    "🎁 Sovrinlar", "🎁 Призы", "🎁 Prizes"
}))
async def on_prizes(message: Message, lang: str = "uz"):
    """🎁 Show prizes info"""
    content, image_id = await get_text_with_image("prizes", lang)

    if image_id:
        await message.answer_photo(photo=image_id, caption=content, parse_mode="Markdown")
    else:
        await message.answer(content, parse_mode="Markdown")


# =============================================
# 👑 ADMIN REPLY KEYBOARD HANDLERS
# These handle the reply keyboard buttons for admins
# =============================================

@router.message(F.text.in_({
    "📊 Statistika", "📊 Статистика", "📊 Statistics"
}))
async def on_admin_stats_reply(message: Message, lang: str = "uz"):
    """📊 Admin stats from reply keyboard"""
    if not await db.is_user_admin(message.from_user.id):
        return
    stats = await db.get_statistics()
    text = await get_message("stats_message", lang, **stats)
    await message.answer(text, parse_mode="Markdown")


@router.message(F.text.in_({
    "📥 Excel yuklab olish", "📥 Скачать Excel", "📥 Download Excel"
}))
async def on_admin_export_reply(message: Message, bot: Bot, lang: str = "uz"):
    """📥 Admin Excel export from reply keyboard"""
    if not await db.is_user_admin(message.from_user.id):
        return

    gen_text = await get_message("export_generating", lang)
    await message.answer(gen_text, parse_mode="Markdown")

    try:
        import io
        from datetime import datetime, timezone, timedelta
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from aiogram.types import BufferedInputFile

        users = await db.get_all_users_for_export()
        wb = Workbook()
        ws = wb.active
        ws.title = "Giveaway Users"

        # 🎨 Styles
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

        headers = [
            "№", "Telegram ID", "Username", "Ism",
            "Familiya", "Telefon", "Til", "Tasdiqlangan",
            "Kodlar", "Referal soni", "Jami kodlar", "Sana"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for idx, user in enumerate(users, 1):
            row = idx + 1
            data = [
                idx, user.get("telegram_id", ""),
                user.get("username", "") or "—",
                user.get("first_name", "") or "—",
                user.get("last_name", "") or "—",
                user.get("phone", "") or "—",
                user.get("language", "uz").upper(),
                "✅" if user.get("is_verified") else "❌",
                user.get("codes_list", "") or "—",
                user.get("referral_count", 0),
                user.get("total_codes", 0),
                user.get("created_at", "").strftime("%d.%m.%Y %H:%M") if user.get("created_at") else "—"
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.alignment = center_align if col in (1, 7, 8, 10, 11) else cell_align
                if idx % 2 == 0:
                    cell.fill = light_fill

        for col in range(1, len(headers) + 1):
            max_len = len(str(headers[col - 1]))
            for r in range(2, len(users) + 2):
                cl = len(str(ws.cell(row=r, column=col).value or ""))
                if cl > max_len:
                    max_len = cl
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        tz = timezone(timedelta(hours=5))
        now = datetime.now(tz)
        filename = f"giveaway_users_{now.strftime('%Y%m%d_%H%M')}.xlsx"

        doc = BufferedInputFile(output.read(), filename=filename)
        done_text = await get_message("export_done", lang)
        await bot.send_document(message.from_user.id, doc, caption=done_text)

    except Exception as e:
        logger.error(f"❌ Excel export error: {e}")
        await message.answer(f"❌ Export error: {e}")


@router.message(F.text.in_({
    "📣 Barchaga xabar yuborish", "📣 Сделать рассылку", "📣 Broadcast message",
    "📢 Broadcast", "📢 Рассылка"
}))
async def on_admin_broadcast_reply(message: Message, state: FSMContext, lang: str = "uz"):
    """📢 Start broadcast from reply keyboard"""
    if not await db.is_user_admin(message.from_user.id):
        return

    from bot.handlers.admin import AdminStates
    text = await get_message("send_broadcast_text", lang)
    await message.answer(text, parse_mode="Markdown")
    await state.set_state(AdminStates.broadcast_compose)


@router.message(F.text.in_({
    "🔄 Kanal sozlamalari", "🔄 Каналы", "🔄 Channel settings"
}))
async def on_admin_channels_reply(message: Message, lang: str = "uz"):
    """🔄 Channel settings from reply keyboard"""
    if not await db.is_user_admin(message.from_user.id):
        return

    from bot.handlers.admin import _build_channels_text_and_keyboard
    channels = await db.get_all_channels()
    text, keyboard = _build_channels_text_and_keyboard(channels)
    await message.answer(text, parse_mode="Markdown", reply_markup=keyboard)


@router.message(F.text.in_({
    "📝 Textlarni o'zgartirish", "📝 Редактировать тексты", "📝 Edit texts"
}))
async def on_admin_edit_texts_reply(message: Message, lang: str = "uz"):
    """📝 Edit texts from reply keyboard"""
    if not await db.is_user_admin(message.from_user.id):
        return

    text = await get_message("select_text_key", lang)
    await message.answer(text, parse_mode="Markdown", reply_markup=get_text_keys_keyboard())
